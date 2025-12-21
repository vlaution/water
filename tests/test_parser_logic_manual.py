import openpyxl
from io import BytesIO
import sys
import os

# Add backend to path
sys.path.append(os.getcwd())

from backend.parser.valuation_parser import ValuationExcelParser

def create_mock_excel():
    wb = openpyxl.Workbook()
    
    # 1. Sheet: Inp_1 (Required for init)
    ws1 = wb.active
    ws1.title = "Inp_1"
    ws1.cell(row=3, column=3, value="Test Corp")
    ws1.cell(row=5, column=3, value="2024-12-31")
    ws1.cell(row=7, column=3, value="USD")
    ws1.cell(row=9, column=3, value=0.25)
    
    # 2. Sheet: Back_end_links (Target logic)
    ws2 = wb.create_sheet("Back_end_links")
    
    # Setup Financials Header at Row 20 (Columns 7-12)
    # Col 7 = G
    ws2.cell(row=20, column=7, value="LTM") # The Critical Value
    ws2.cell(row=20, column=8, value=2024)
    
    # Data Row 21 (Revenue)
    ws2.cell(row=21, column=7, value=1000)
    ws2.cell(row=21, column=8, value=1100)
    
    # Data Row 23 (EBITDA)
    ws2.cell(row=23, column=7, value=200)
    ws2.cell(row=23, column=8, value=250)
    
    # Setup Valuation Results (Dynamic Row Search Test)
    # Put header at Row 5 to test dynamic search (default was 3)
    ws2.cell(row=5, column=1, value="Valuation Approach")
    ws2.cell(row=5, column=2, value="Method")
    ws2.cell(row=5, column=3, value="Enterprise Value")
    
    ws2.cell(row=6, column=1, value="Market Approach")
    ws2.cell(row=6, column=2, value="Test Method")
    ws2.cell(row=6, column=3, value=5000)
    ws2.cell(row=6, column=4, value=1.0) # Weight

    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    return virtual_workbook.getvalue()

def test_manual_scan():
    print("Creating mock Excel...")
    content = create_mock_excel()
    
    print("Initializing Parser...")
    parser = ValuationExcelParser(content)
    
    print("Parsing...")
    data = parser.parse()
    
    # Verify Financials
    print("\n--- Financials Verification ---")
    ltm_found = False
    for fin in data.financials:
        print(f"Year: {fin.year} (Type: {type(fin.year)}) | Revenue: {fin.revenue}")
        if str(fin.year) == "LTM" and fin.revenue == 1000:
            ltm_found = True
            
    if ltm_found:
        print("SUCCESS: LTM Financials found!")
    else:
        print("FAILURE: LTM Financials NOT found.")
        
    # Verify Results
    print("\n--- Valuation Results Verification ---")
    results = data.valuation_results
    if results and results[0].method == "Test Method":
        print(f"SUCCESS: Found method '{results[0].method}' at dynamic row.")
    else:
        print(f"FAILURE: Dynamic row search failed. Found: {results}")

if __name__ == "__main__":
    test_manual_scan()
