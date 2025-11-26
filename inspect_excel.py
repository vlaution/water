import openpyxl
import pandas as pd
import sys

FILE_PATH = "c:/Users/Abhiram/Desktop/water/valuation_automation_dashboard.xlsm"

def inspect_excel(file_path):
    try:
        print(f"Loading {file_path}...")
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        print(f"Sheet names: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            print(f"\n--- Sheet: {sheet_name} ---")
            ws = wb[sheet_name]
            # Print first 5 rows to identify headers
            for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
                print(f"Row {i+1}: {row}")
                
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    inspect_excel(FILE_PATH)
