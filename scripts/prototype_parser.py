import openpyxl
import json
from datetime import datetime

class GracefulParser:
    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(file_path, data_only=True)
        self.sheet = self.wb['Inp_1']

    def extract_value(self, row_idx, col_idx=3):
        """Extracts value from a specific row and column (default column C=3)"""
        return self.sheet.cell(row=row_idx, column=col_idx).value

    def parse(self):
        # Corrected mapping based on DataFrame index to Excel row conversion
        # Header is Row 1. Index 0 is Row 2. Index 1 is Row 3.
        data = {
            "company_name": self.extract_value(3),   # Row 3
            "valuation_date": str(self.extract_value(5)), # Row 5
            "currency": self.extract_value(7),       # Row 7
            "tax_rate": self.extract_value(9),       # Row 9
            "geography": self.parse_geography()
        }
        return data

    def parse_geography(self):
        # Geography details start at Row 12 (Global) to Row 20 (Others)
        geo_map = {}
        for row in range(12, 21):
            label = self.sheet.cell(row=row, column=2).value
            is_active = self.sheet.cell(row=row, column=3).value
            if label:
                geo_map[label] = bool(is_active)
        return geo_map

if __name__ == "__main__":
    file_path = r"C:\Users\Abhiram\Desktop\water\Valuation_Automation_Dashboard_Draft 12_AJ.xlsm"
    try:
        parser = GracefulParser(file_path)
        result = parser.parse()
        print(json.dumps(result, indent=4))
    except Exception as e:
        print(f"Error: {e}")
