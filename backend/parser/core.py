import openpyxl
from typing import Dict, Any, List
from .models import WorkbookData, SheetData
from .utils import clean_header

class SheetParser:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb = None

    def load(self):
        self.wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)

    def parse_sheet(self, sheet_name: str) -> SheetData:
        if not self.wb:
            raise ValueError("Workbook not loaded. Call load() first.")
            
        ws = self.wb[sheet_name]
        # Use values_only=True for performance
        rows = list(ws.iter_rows(values_only=True))
        
        if not rows:
            return SheetData(name=sheet_name, headers=[], rows=[])

        # Heuristic: Find the header row.
        # We look for the first row that has a significant number of string cells.
        header_row_idx = 0
        headers = []
        
        found_header = False
        for i, row in enumerate(rows):
            # Skip empty rows
            if not any(row):
                continue
                
            # Count string cells
            str_count = sum(1 for cell in row if isinstance(cell, str) and cell.strip())
            non_empty_count = sum(1 for cell in row if cell is not None)
            
            # If > 50% of non-empty cells are strings, assume it's a header
            if non_empty_count > 0 and (str_count / non_empty_count) > 0.5:
                header_row_idx = i
                headers = [clean_header(cell) for cell in row]
                found_header = True
                break
        
        if not found_header:
            # Fallback to first row if no obvious header found
            header_row_idx = 0
            headers = [clean_header(cell) for cell in rows[0]]

        data_rows = []
        # Iterate over subsequent rows
        for row in rows[header_row_idx+1:]:
            # Skip completely empty rows
            if not any(row):
                continue
                
            row_dict = {}
            has_data = False
            for i, cell in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_dict[headers[i]] = cell
                    has_data = True
            
            if has_data:
                data_rows.append(row_dict)
            
        return SheetData(name=sheet_name, headers=headers, rows=data_rows)

    def parse(self) -> WorkbookData:
        if not self.wb:
            self.load()
            
        sheets_data = {}
        for sheet_name in self.wb.sheetnames:
            # Skip hidden or temporary sheets if needed, but user said "Treat each sheet as an independent module"
            # So we parse everything.
            try:
                sheets_data[sheet_name] = self.parse_sheet(sheet_name)
            except Exception as e:
                print(f"Error parsing sheet {sheet_name}: {e}")
                # Continue with other sheets
                continue
            
        return WorkbookData(sheets=sheets_data)
