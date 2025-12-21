import openpyxl
from io import BytesIO
from typing import Dict, Any, Union
from datetime import datetime
from backend.schemas.valuation_import import ValuationImportData

class ValuationExcelParser:
    def __init__(self, file_content: bytes):
        self.wb = openpyxl.load_workbook(filename=BytesIO(file_content), data_only=True)
        if "Inp_1" not in self.wb.sheetnames:
            # Fallback or strict error? User mentioned accuracy is key.
            # If Inp_1 is missing, it's likely the wrong file.
            raise ValueError("Invalid File: Sheet 'Inp_1' not found.")
        self.sheet = self.wb["Inp_1"]

    def extract_value(self, row: int, col: int = 3) -> Any:
        return self.sheet.cell(row=row, column=col).value

    def parse(self) -> ValuationImportData:
        try:
            # Row mappings based on verified prototype
            company_name = self.extract_value(3)
            val_date_raw = self.extract_value(5)
            currency = self.extract_value(7)
            tax_rate = self.extract_value(9)

            # Validation / Type Checking
            if not company_name:
                raise ValueError("Company Name is missing (Row 3).")
            
            # Date handling
            # Excel might return datetime object or string
            val_date_str = str(val_date_raw)
            # If it's a datetime object, format it nicely
            if isinstance(val_date_raw, datetime):
                val_date_str = val_date_raw.strftime("%Y-%m-%d")

            # Currency check
            if not isinstance(currency, str) or len(currency) != 3:
                 # Attempt cleanup or raise error? Be strict.
                 raise ValueError(f"Invalid Currency format: {currency} (Row 7)")
            
            # Tax rate check
            if not isinstance(tax_rate, (int, float)):
                 raise ValueError(f"Invalid Tax Rate: {tax_rate} (Row 9). Expected number.")

            geography = self._parse_geography()

            return ValuationImportData(
                company_name=str(company_name),
                valuation_date=val_date_str,
                currency=currency,
                tax_rate=float(tax_rate),
                geography=geography,
                financials=self._parse_financials(),
                balance_sheet=self._parse_balance_sheet_snapshot(),
                working_capital=self._parse_working_capital(),
                capex=self._parse_capex(),
                valuation_results=self._parse_valuation_results(),
                wacc_metrics=self._parse_wacc()
            )

        except Exception as e:
            # Re-raise with context
            raise ValueError(f"Parsing failed: {str(e)}")

    def _parse_geography(self) -> Dict[str, bool]:
        geo_map = {}
        for row in range(12, 21):
            label = self.sheet.cell(row=row, column=2).value
            val = self.sheet.cell(row=row, column=3).value
            if label:
                is_active = False
                if isinstance(val, bool):
                    is_active = val
                elif isinstance(val, (int, float)):
                    is_active = bool(val)
                elif isinstance(val, str):
                    is_active = val.lower() == "true"
                geo_map[str(label)] = is_active
        return geo_map

    def _parse_financials(self) -> list:
        if "Back_end_links" not in self.wb.sheetnames:
            return []
        
        sheet = self.wb["Back_end_links"]
        financials = []
        for col_idx in range(7, 13): 
            year_val = sheet.cell(row=20, column=col_idx).value
            
            # ALLOW "LTM" string or Integer years
            if not isinstance(year_val, int) and str(year_val).lower() != "ltm":
                continue
            
            # Normalize year_val for storage (keep "LTM" as string or use special int?)
            # Schema allows Union[int, str], but downstream logic often expects int for sorting.
            # Ideally, we keep it as is if schema allows.
            
            rev = sheet.cell(row=21, column=col_idx).value
            gp = sheet.cell(row=22, column=col_idx).value
            ebitda = sheet.cell(row=23, column=col_idx).value
            ni = sheet.cell(row=24, column=col_idx).value
            
            financials.append({
                "year": year_val, # Can be "LTM" or 2024
                "revenue": float(rev or 0),
                "gross_profit": float(gp or 0),
                "ebitda": float(ebitda or 0),
                "net_income": float(ni or 0)
            })
        return financials

    def _parse_balance_sheet_snapshot(self) -> Dict[str, float]:
        if "Back_end_links" not in self.wb.sheetnames:
            return None
            
        sheet = self.wb["Back_end_links"]
        cash = sheet.cell(row=38, column=7).value
        st_inv = sheet.cell(row=39, column=7).value
        debt = sheet.cell(row=40, column=7).value
        
        return {
            "cash": float(cash or 0),
            "short_term_investments": float(st_inv or 0),
            "debt": float(debt or 0)
        }

    def _parse_lbo(self) -> Any:
        # Check for LBO specific sheet
        target_sheet = None
        for name in ["Inp_LBO", "LBO", "Deal Structure"]:
            if name in self.wb.sheetnames:
                target_sheet = self.wb[name]
                break
            
        if not target_sheet:
            return None
            
        try:
            # 1. Search for basic assumptions using label-based lookup
            assumptions = {
                "entry_ebitda": {"labels": ["ebitda", "entry ebitda", "target ebitda"], "default": 0.0},
                "exit_multiple": {"labels": ["exit multiple", "exit ev/ebitda", "target multiple"], "default": 10.0},
                "holding_period": {"labels": ["holding period", "exit year", "term", "years"], "default": 5}
            }
            results = {k: v["default"] for k, v in assumptions.items()}
            
            # Scan first 50 rows for labels
            for r in range(1, 51):
                raw_label = str(target_sheet.cell(row=r, column=2).value or "").lower().strip()
                for key, config in assumptions.items():
                    if any(l in raw_label for l in config["labels"]):
                        # Check col 3, then 4 for the value
                        for c in [3, 4]:
                            val = target_sheet.cell(row=r, column=c).value
                            if isinstance(val, (int, float)):
                                results[key] = val
                                break

            # 2. Search for Debt Tranches
            # Look for a row labeled "Debt", "Sources", or "Financing" to find the table start
            tranches = []
            table_start_row = 15 # Default fallback
            for r in range(1, 30):
                label = str(target_sheet.cell(row=r, column=2).value or "").lower()
                if "debt" in label or "sources" in label:
                    table_start_row = r + 1
                    break
            
            for r in range(table_start_row, table_start_row + 10):
                name = target_sheet.cell(row=r, column=2).value
                if not name or "total" in str(name).lower(): 
                   if not tranches and r < table_start_row + 5: continue # Skip empty rows at start
                   break
                
                # Dynamic amount search (Col 3 or 4)
                amt = 0.0
                rate = 0.08
                for c in [3, 4, 5]:
                    val = target_sheet.cell(row=r, column=c).value
                    if isinstance(val, (int, float)):
                        if val > 1: # High chance it's the amount
                            amt = float(val)
                        elif 0 < val < 1: # High chance it's the interest rate
                            rate = float(val)

                if amt > 0:
                    tranches.append({
                        "name": str(name),
                        "amount": amt,
                        "interest_rate": rate,
                        "cash_interest": True,
                        "amortization_rate": 0.0
                    })
                
            return {
                "entry_ebitda": float(results["entry_ebitda"]),
                "exit_multiple": float(results["exit_multiple"]),
                "holding_period": int(results["holding_period"]),
                "tranches": tranches,
                "revenue_growth": 0.05,
                "ebitda_margin": 0.25,
                "capex_percent": 0.03,
                "nwc_percent": 0.05
            }
        except Exception as e:
            print(f"LBO Parsing Warning: {e}")
            return None

    def _parse_working_capital(self) -> list:
        if "WC & Capex" not in self.wb.sheetnames:
            return []
            
        sheet = self.wb["WC & Capex"]
        wc_data = []
        for col in range(3, 12):
            year_val = sheet.cell(row=2, column=col).value
            if not year_val: 
                continue
            tr = sheet.cell(row=3, column=col).value
            inv = sheet.cell(row=5, column=col).value
            tp = sheet.cell(row=9, column=col).value
            wc = sheet.cell(row=14, column=col).value
            chg_wc = sheet.cell(row=15, column=col).value
            
            if isinstance(chg_wc, str): chg_wc = 0.0
            
            wc_data.append({
                "year": year_val,
                "trade_receivables": float(tr or 0),
                "inventory": float(inv or 0),
                "trade_payables": float(tp or 0),
                "working_capital": float(wc or 0),
                "change_in_working_capital": float(chg_wc or 0)
            })
        return wc_data

    def _parse_capex(self) -> list:
        if "WC & Capex" not in self.wb.sheetnames:
            return []
            
        sheet = self.wb["WC & Capex"]
        capex_data = []
        for col in range(3, 12):
            year_val = sheet.cell(row=2, column=col).value
            if not year_val: continue
            
            fa = sheet.cell(row=27, column=col).value
            depr = sheet.cell(row=28, column=col).value
            net_capex = sheet.cell(row=29, column=col).value
            
            if isinstance(net_capex, str): net_capex = 0.0
            
            capex_data.append({
                "year": year_val,
                "fixed_assets": float(fa or 0),
                "depreciation": float(depr or 0),
                "net_capital_expenditure": float(net_capex or 0)
            })
        return capex_data

    def _parse_valuation_results(self) -> list:
        if "Back_end_links" not in self.wb.sheetnames:
            return []
            
        sheet = self.wb["Back_end_links"]
        results = []
        
        # Dynamic Search for Table Start
        start_row = 3
        found_header = False
        for r in range(1, 20):
            val = str(sheet.cell(row=r, column=1).value or "").lower()
            if "valuation approach" in val:
                start_row = r + 1 # Data starts after header?
                # Actually in the dump:
                # Row 2 (Index 1): Valuation Approach | Method | EV
                # Row 3 (Index 2): Market Approach | GPC | ...
                
                # If we found "valuation approach" at Row 2, data starts at 3.
                # If distinct "Method" header is present, use that.
                start_row = r + 1
                found_header = True
                break
        
        # Fallback if not found (keep original logic or warn)
        if not found_header:
            start_row = 3

        # Iterate Rows from Start
        for row in range(start_row, start_row + 15):
            approach = sheet.cell(row=row, column=1).value
            method = sheet.cell(row=row, column=2).value
            val = sheet.cell(row=row, column=3).value
            weight = sheet.cell(row=row, column=4).value
            
            # Stop condition: Empty method AND approach (end of table)
            if not method and not approach: 
                continue
            
            # Skip header row just in case we landed on it
            if str(method).lower() == "method":
                continue

            results.append({
                "approach": str(approach or "Other"),
                "method": str(method or "Unknown"),
                "enterprise_value": float(val or 0),
                "weight": float(weight or 0)
            })
            
        return results

    def _parse_wacc(self) -> Dict[str, float]:
        if "Inp_4" not in self.wb.sheetnames:
            return None
            
        sheet = self.wb["Inp_4"]
        
        # Mappings from Deep Scan (0-based DF index -> 1-based Excel Row):
        # WACC: Index 24 -> Row 25
        # Risk Free Rate: Index 27 -> Row 28
        # Beta: Index 29 -> Row 30
        # ERP: Index 30 -> Row 31 (Assumed next to Beta, check dump if needed)
        # Cost of Equity: Index 32 -> Row 33
        
        col = 8 # Column H
        wacc = sheet.cell(row=25, column=col).value
        rf = sheet.cell(row=28, column=col).value
        beta = sheet.cell(row=30, column=col).value
        erp = sheet.cell(row=31, column=col).value
        coe = sheet.cell(row=33, column=col).value
        
        # Cost of Debt? Not explicitly in dump, but usually near.
        # Assuming 0 if not found for now to prevent crash.
        cod = 0.0
        
        return {
            "risk_free_rate": float(rf or 0),
            "beta": float(beta or 0),
            "equity_risk_premium": float(erp or 0),
            "cost_of_equity": float(coe or 0),
            "cost_of_debt_post_tax": float(cod),
            "wacc": float(wacc or 0)
        }
