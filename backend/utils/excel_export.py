from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from io import BytesIO
import json

def create_valuation_excel(run_data: dict) -> BytesIO:
    """
    Create an Excel workbook with valuation results
    """
    # Parse JSON strings if needed
    if isinstance(run_data.get('input_data'), str):
        run_data['input_data'] = json.loads(run_data['input_data'])
    if isinstance(run_data.get('results'), str):
        run_data['results'] = json.loads(run_data['results'])
    
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create Summary sheet
    ws_summary = wb.create_sheet("Summary")
    create_summary_sheet(ws_summary, run_data)
    
    # Create Inputs sheet
    ws_inputs = wb.create_sheet("Inputs")
    create_inputs_sheet(ws_inputs, run_data)
    
    # Create Results sheet
    ws_results = wb.create_sheet("Results")
    create_results_sheet(ws_results, run_data)
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file

def create_summary_sheet(ws, run_data):
    """Create summary sheet with key metrics"""
    # Header styling
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14)
    
    # Title
    ws['A1'] = "Valuation Summary"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:B1')
    
    # Company info
    ws['A3'] = "Company:"
    ws['B3'] = run_data.get('company_name', 'N/A')
    ws['A4'] = "Date:"
    ws['B4'] = run_data.get('created_at', 'N/A')
    ws['A5'] = "Mode:"
    ws['B5'] = run_data.get('mode', 'N/A')
    
    # Key metrics
    results = run_data.get('results', {})
    ws['A7'] = "Key Metrics"
    ws['A7'].font = header_font
    ws['A7'].fill = header_fill
    ws['B7'].fill = header_fill
    
    ws['A8'] = "Enterprise Value"
    ws['B8'] = results.get('enterprise_value', 0)
    ws['B8'].number_format = '$#,##0'
    
    ws['A9'] = "Equity Value"
    ws['B9'] = results.get('equity_value', 0)
    ws['B9'].number_format = '$#,##0'
    
    ws['A10'] = "WACC"
    ws['B10'] = results.get('wacc', 0)
    ws['B10'].number_format = '0.00%'
    
    # Valuation methods
    methods = results.get('methods', {})
    ws['A12'] = "Valuation Methods"
    ws['A12'].font = header_font
    ws['A12'].fill = header_fill
    ws['B12'].fill = header_fill
    ws['C12'].fill = header_fill
    
    ws['A13'] = "Method"
    ws['B13'] = "Value"
    ws['C13'] = "Weight"
    
    row = 14
    for method_name, method_data in methods.items():
        ws[f'A{row}'] = method_name
        ws[f'B{row}'] = method_data.get('value', 0)
        ws[f'B{row}'].number_format = '$#,##0'
        ws[f'C{row}'] = method_data.get('weight', 0)
        ws[f'C{row}'].number_format = '0.00%'
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15

def create_inputs_sheet(ws, run_data):
    """Create inputs sheet with all input data"""
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    ws['A1'] = "Valuation Inputs"
    ws['A1'].font = Font(bold=True, size=16)
    
    input_data = run_data.get('input_data', {})
    
    # DCF Inputs
    dcf_input = input_data.get('dcf_input', {})
    if dcf_input:
        ws['A3'] = "DCF Inputs"
        ws['A3'].font = header_font
        ws['A3'].fill = header_fill
        
        # Historical data
        historical = dcf_input.get('historical', {})
        if historical:
            ws['A5'] = "Historical Financials"
            ws['A5'].font = Font(bold=True)
            
            years = historical.get('years', [])
            row = 6
            for i, year in enumerate(years, start=2):
                ws.cell(row=row, column=i).value = year
            
            metrics = ['revenue', 'ebitda', 'ebit', 'net_income', 'capex', 'nwc']
            row = 7
            for metric in metrics:
                ws.cell(row=row, column=1).value = metric.upper()
                values = historical.get(metric, [])
                for i, value in enumerate(values, start=2):
                    cell = ws.cell(row=row, column=i)
                    cell.value = value
                    cell.number_format = '#,##0'
                row += 1
        
        # Projections
        projections = dcf_input.get('projections', {})
        if projections:
            ws['A15'] = "Projection Assumptions"
            ws['A15'].font = Font(bold=True)
            
            row = 16
            for key, value in projections.items():
                ws.cell(row=row, column=1).value = key.replace('_', ' ').title()
                cell = ws.cell(row=row, column=2)
                cell.value = value
                if 'rate' in key or 'margin' in key:
                    cell.number_format = '0.00%'
                row += 1
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15

def create_results_sheet(ws, run_data):
    """Create detailed results sheet"""
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    ws['A1'] = "Detailed Results"
    ws['A1'].font = Font(bold=True, size=16)
    
    results = run_data.get('results', {})
    
    # Write all results as key-value pairs
    row = 3
    for key, value in results.items():
        if key not in ['input_summary', 'methods']:  # Skip nested objects
            ws.cell(row=row, column=1).value = key.replace('_', ' ').title()
            ws.cell(row=row, column=2).value = value
            if isinstance(value, (int, float)) and value > 1000:
                ws.cell(row=row, column=2).number_format = '#,##0'
            row += 1
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
